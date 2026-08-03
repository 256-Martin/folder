"""
Extract every populated row from the Ugabrush Google Sheets export into a single
JSON document used to seed the web application.

Run:  python scripts/extract_seed.py
Out:  src/db/seed-data.json
"""

import json
import os
import sys
import datetime as dt

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  python -m pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEFAULT_XLSX = os.path.expanduser(r"~\Downloads\Ugabrush Manufacturing System V1.7.xlsx")
XLSX = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
OUT = os.path.join(ROOT, "src", "db", "seed-data.json")

if not os.path.exists(XLSX):
    sys.exit(f"Workbook not found: {XLSX}\nPass the path as the first argument.")

wb = openpyxl.load_workbook(XLSX, data_only=True)


def norm(v):
    """Normalise a cell value into something JSON-serialisable."""
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        # Keep the time component; business-date fields slice the first 10 chars.
        return v.isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def rows(sheet_name, cols, start=4, key_col=None, limit=1200):
    """Read `sheet_name` from `start` downwards, mapping column letters to names.

    `cols` maps output field name -> column letter.
    A row is kept when `key_col` (default: first mapped column) is non-empty.
    """
    ws = wb[sheet_name]
    kc = key_col or next(iter(cols.values()))
    out = []
    for r in range(start, min(ws.max_row, start + limit) + 1):
        if norm(ws[f"{kc}{r}"].value) is None:
            continue
        rec = {name: norm(ws[f"{letter}{r}"].value) for name, letter in cols.items()}
        rec["_row"] = r
        out.append(rec)
    return out


def keyvalue(sheet_name, key="A", value="B", note="C", start=4):
    ws = wb[sheet_name]
    out = []
    for r in range(start, ws.max_row + 1):
        k = norm(ws[f"{key}{r}"].value)
        if k is None:
            continue
        out.append(
            {
                "key": k,
                "value": norm(ws[f"{value}{r}"].value),
                "note": norm(ws[f"{note}{r}"].value),
            }
        )
    return out


data = {}

# ----------------------------------------------------------------- masters ---
data["items"] = rows(
    "Item Master",
    {
        "code": "A",
        "name": "B",
        "category": "C",
        "baseUom": "D",
        "standardCost": "E",
        "reorderLevel": "F",
        "trackedByBatch": "G",
        "active": "H",
    },
)

data["suppliers"] = rows(
    "Supplier Master",
    {
        "code": "A",
        "name": "B",
        "categorySupplied": "C",
        "contact": "D",
        "active": "E",
        "supplierType": "F",
    },
)

data["workers"] = rows(
    "Worker Master",
    {"code": "A", "name": "B", "processesAbleToDo": "C", "active": "D"},
)

data["processes"] = rows(
    "Process Master",
    {
        "code": "A",
        "name": "B",
        "sequenceNo": "C",
        "inputStage": "D",
        "outputStage": "E",
        "consumesMaterials": "F",
        "producesStockItem": "G",
    },
)

data["workerSkills"] = rows(
    "Worker Process Skills",
    {
        "workerCode": "A",
        "processCode": "C",
        "skillStatus": "E",
        "active": "F",
        "notes": "G",
    },
)

data["pieceRates"] = rows(
    "Piece Rate Settings",
    {
        "processCode": "B",
        "productType": "C",
        "ratePerUnit": "F",
        "unit": "G",
        "effectiveFrom": "H",
        "active": "I",
        "notes": "K",
    },
    key_col="B",
)

# ------------------------------------------------------------ transactions ---
data["purchases"] = rows(
    "Purchases Log",
    {
        "date": "B",
        "supplier": "C",
        "itemCode": "D",
        "batchNo": "F",
        "qty": "G",
        "totalCost": "I",
        "qualityNotes": "K",
    },
    key_col="B",
)

data["movements"] = rows(
    "Inventory Ledger",
    {
        "date": "B",
        "itemCode": "C",
        "batchNo": "E",
        "movementType": "F",
        "qty": "G",
        "unitCost": "I",
        "refSource": "L",
        "by": "M",
        "note": "N",
        "issuedToType": "Q",
        "issuedTo": "R",
        "receivedBy": "S",
    },
    key_col="B",
)

data["batches"] = rows(
    "Batch Register",
    {
        "batchNo": "A",
        "itemCode": "B",
        "supplier": "D",
        "purchaseDate": "E",
        "qtyReceived": "F",
        "unitCost": "H",
        "quality": "M",
    },
)

data["operations"] = rows(
    "Production Operations Log",
    {
        "date": "B",
        "worker": "C",
        "processCode": "D",
        "productType": "E",
        "inputBatch": "F",
        "outputBatch": "G",
        "acceptedQty": "H",
        "rejectedQty": "I",
        "rejectReason": "J",
        "notes": "O",
    },
    key_col="B",
)

data["dispatches"] = rows(
    "Sales - Dispatch Log",
    {
        "date": "B",
        "destinationName": "C",
        "productCode": "D",
        "qty": "F",
        "unitPrice": "G",
        "note": "I",
        "destinationType": "L",
        "personResponsible": "M",
        "deliveryNoteNo": "N",
        "salesOrderNo": "O",
    },
    key_col="B",
)

data["meals"] = rows(
    "Meal Log",
    {
        "date": "B",
        "worker": "C",
        "plateCount": "D",
        "qualified": "E",
        "qualifiedPlates": "F",
        "unqualifiedPlates": "G",
        "companyContribution": "H",
        "workerTopUp": "I",
        "fullCostDeduction": "J",
        "note": "L",
        "approvedBy": "M",
        "companyFullySponsored": "Q",
        "workerCashPaid": "R",
        "foodProvider": "T",
        "actualPlateCost": "U",
        "actualCompanyContribution": "V",
        "workerRequiredContribution": "W",
        "supplierPriceChanged": "X",
        "contributionChanged": "Y",
        "reasonForChange": "Z",
    },
    key_col="B",
)

data["deductions"] = rows(
    "Deductions Log",
    {
        "date": "B",
        "worker": "C",
        "deductionType": "D",
        "amount": "E",
        "reason": "F",
        "approvedBy": "G",
        "status": "H",
        "notes": "I",
    },
    key_col="B",
)

data["expenses"] = rows(
    "Expense & Provider Payments",
    {
        "date": "B",
        "provider": "C",
        "expenseCategory": "E",
        "periodFrom": "F",
        "periodTo": "G",
        "plates": "H",
        "plateCost": "I",
        "totalBill": "J",
        "amountPaid": "K",
        "accountPaidFrom": "M",
        "paymentMethod": "N",
        "transactionNo": "O",
        "paidBy": "P",
        "notes": "Q",
    },
    key_col="B",
)

# ------------------------------------------------------------- meal engine ---
data["mealRules"] = rows(
    "Meal Qualification Settings",
    {
        "ruleId": "A",
        "processCode": "B",
        "ruleBasis": "C",
        "product": "D",
        "condition": "E",
        "threshold": "F",
        "weeklySmallMin": "G",
        "weeklyLargeMin": "H",
        "eatingDaysEarned": "I",
        "active": "J",
        "notes": "K",
    },
)

data["mealCostSettings"] = keyvalue("Meal Cost Settings")

# ----------------------------------------------------------------- config ----
data["settings"] = keyvalue("Settings")
data["reportHeader"] = keyvalue("Report Header Settings")

data["accountingMappings"] = rows(
    "Accounting Export Settings",
    {"mappingKey": "A", "debitAccount": "B", "creditAccount": "C", "notes": "D"},
)

# Lists: each pair of columns is (label, values...). Capture every column that
# has a header in row 3 so dropdowns match the sheet exactly.
lists_ws = wb["Lists"]
lists = {}
for c in range(1, lists_ws.max_column + 1):
    header = norm(lists_ws.cell(row=3, column=c).value)
    if not header:
        continue
    values = []
    for r in range(4, lists_ws.max_row + 1):
        v = norm(lists_ws.cell(row=r, column=c).value)
        if v is not None:
            values.append(v)
    if values:
        lists[header] = values
data["lists"] = lists

# ----------------------------------------------------------------- system ----
data["auditLog"] = rows(
    "Audit Log",
    {
        "timestamp": "A",
        "user": "B",
        "action": "C",
        "sheet": "D",
        "ref": "E",
        "details": "F",
        "result": "G",
    },
    key_col="A",
)

data["reportLog"] = rows(
    "Report Log",
    {
        "timestamp": "A",
        "user": "B",
        "action": "C",
        "reportType": "D",
        "period": "E",
        "filters": "F",
        "note": "G",
        "status": "H",
        "output": "I",
    },
    key_col="A",
)

data["voidRegister"] = rows(
    "Void Register",
    {
        "timestamp": "B",
        "logType": "C",
        "sheet": "D",
        "sourceRow": "E",
        "entryId": "F",
        "reason": "G",
        "voidedBy": "H",
        "status": "I",
        "restoredAt": "J",
        "restoredBy": "K",
        "reversalEffect": "L",
        "oldValues": "M",
    },
    key_col="B",
)

data["batchRenameLog"] = rows(
    "Batch Rename Log",
    {
        "timestamp": "A",
        "oldBatchNo": "B",
        "newBatchNo": "C",
        "item": "D",
        "oldPurchaseDate": "E",
        "newPurchaseDate": "F",
        "changedBy": "G",
        "reason": "H",
        "sourcePurchaseId": "I",
        "status": "J",
    },
    key_col="A",
)

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as fh:
    json.dump(data, fh, indent=1, ensure_ascii=False)

print(f"Wrote {OUT}")
for k, v in data.items():
    n = len(v) if isinstance(v, (list, dict)) else 1
    print(f"  {k:22} {n}")
